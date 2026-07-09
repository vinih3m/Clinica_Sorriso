from django.views.generic import View ,TemplateView, CreateView, ListView, UpdateView, DeleteView
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.contrib.auth.models import User, Group
from django.contrib.auth.forms import PasswordChangeForm
from django.contrib.auth import update_session_auth_hash
from django.shortcuts import redirect,  render, get_object_or_404
from django.urls import reverse
from openpyxl.styles import Font, Alignment, PatternFill
from django.contrib import messages
from django.http import JsonResponse , HttpResponse
from django.contrib.messages.views import SuccessMessageMixin
import json, openpyxl
from django.contrib.auth.views import LoginView
from django.db import transaction
from django.forms import modelformset_factory
from .mixins import AvisoPermissaoMixin
from django.contrib.auth.decorators import login_required
from django.utils.decorators import method_decorator
from django.views.decorators.cache import never_cache
from django.urls import reverse_lazy
from .forms import CriarUsuarioForm, UsuarioContaForm, ProfileForm, FotoProfileForm , DentistaProfileForm, DadosDaClinicaForm, ProdutoEstoqueForm, BaseProdutoEstoqueFormSet
from .models import Profile, DentistaProfile, DadosDaClinica, ProdutoEstoque, RetiradaEstoque
from .forms import AgendamentoForm, PacienteForm
from django.template.loader import render_to_string
from django.contrib.auth.views import LoginView
from django.conf import settings
from django.shortcuts import redirect
from django.urls import reverse
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.decorators.http import require_POST
from django.http import JsonResponse
from django.shortcuts import get_object_or_404
from .models import Agendamento
from django.shortcuts import render
from datetime import date, timedelta
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from .forms import AgendamentoForm
from django.views.decorators.http import require_POST
from .models import (Agendamento,Paciente,Profissional,CompromissoAgenda,Etiqueta,EncaixeAgenda,AlertaRetorno,AlertaRetorno,)
from django.http import JsonResponse
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_POST
from django.core.exceptions import ValidationError


# =========================
# LOGIN
# =========================
class CustomLoginView(LoginView):
    template_name = 'registration/login.html'
    redirect_authenticated_user = True

    def get_success_url(self):
        return reverse_lazy('redirect_pos_login')


# =========================
# REDIRECIONAMENTO PÓS LOGIN
# =========================
@login_required(login_url='/login/')
def redirect_pos_login(request):
    return redirect('Tela_Inicial')


# =========================
# ALTERAR SENHA
# =========================
@login_required(login_url='/login/')
def alterar_senha(request):
    if request.method == 'POST':
        form = PasswordChangeForm(user=request.user, data=request.POST)

        if form.is_valid():
            user = form.save()
            update_session_auth_hash(request, user)
            messages.success(request, 'Senha alterada com sucesso!')
        else:
            messages.error(request, 'Não foi possível alterar a senha. Verifique os dados.')

    return redirect('Minha_Conta')

class Dados_Da_ClinicaView(LoginRequiredMixin, UserPassesTestMixin, View):
    pass



# =========================
# TELA INICIAL
# =========================
@method_decorator(never_cache, name='dispatch')
class Tela_InicialView(LoginRequiredMixin, TemplateView):
    template_name = 'Clinica_Sorriso/Tela_Inicial/Tela_Inicial.html'
    login_url = '/login/'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = self.request.user

        context['is_admin'] = user.groups.filter(name='Administrador').exists()
        context['is_dentista'] = user.groups.filter(name='Dentista').exists()
        context['is_recepcionista'] = user.groups.filter(name='Recepcionista').exists()

        return context


# CADASTRAR USUÁRIO
@method_decorator([login_required, never_cache], name='dispatch')
class Criar_UsuarioView(LoginRequiredMixin, AvisoPermissaoMixin, CreateView):
    model = User
    form_class = CriarUsuarioForm
    template_name = 'Clinica_Sorriso/Administrador/Usuario/Cadastrar_Usuario.html'
    success_url = reverse_lazy('Tela_Inicial_Administrador')

    def test_func(self):
        return self.request.user.groups.filter(name='Administrador').exists()

    def form_valid(self, form):
        user = form.save(commit=False)
        user.set_password(form.cleaned_data['password1'])
        user.save()

        grupo_nome = form.cleaned_data['grupo']
        grupo = Group.objects.get(name=grupo_nome)
        user.groups.add(grupo)

        return super().form_valid(form)

# Administrador
@method_decorator(never_cache, name='dispatch')
class Minha_ContaView(LoginRequiredMixin, UserPassesTestMixin, TemplateView):
    template_name = 'Clinica_Sorriso/Minha_Conta/Minha_Conta.html'

    def test_func(self):
        return self.request.user.groups.filter( name__in=['Administrador', 'Dentista' , 'Recepcionista']).exists()

    def handle_no_permission(self):
        return redirect('redirect_pos_login')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        profile, _ = Profile.objects.get_or_create(user=self.request.user)
        dentista, _ = DentistaProfile.objects.get_or_create(profile=profile)

        context['profile'] = profile
        context['dentista'] = dentista
        context['form'] = PasswordChangeForm(user=self.request.user)

        return context

@method_decorator(never_cache, name='dispatch')
class Editar_Minha_ContaView(LoginRequiredMixin, UserPassesTestMixin, View):
    template_name = 'Clinica_Sorriso/Minha_Conta/Editar_Minha_Conta.html'

    def test_func(self): 
        return self.request.user.groups.filter( name__in=['Administrador', 'Dentista', 'Recepcionista']).exists()

    def handle_no_permission(self):
        return redirect('redirect_pos_login')

    def get(self, request):
        profile, _ = Profile.objects.get_or_create(user=request.user)

        user_form = UsuarioContaForm(instance=request.user)
        profile_form = ProfileForm(instance=profile)
        foto_form = FotoProfileForm(instance=profile)

        dentista_form = DentistaProfileForm(
            instance=getattr(profile, 'dentistaprofile', None)
        )

        return render(request, self.template_name, {
            'user_form': user_form,
            'profile_form': profile_form,
            'foto_form': foto_form,
            'dentista_form': dentista_form
        })

    def post(self, request):
        profile, _ = Profile.objects.get_or_create(user=request.user)

        user_form = UsuarioContaForm(request.POST, instance=request.user)
        profile_form = ProfileForm(request.POST, instance=profile)
        foto_form = FotoProfileForm(request.POST, request.FILES, instance=profile)

        dentista_form = DentistaProfileForm(
            request.POST,
            instance=getattr(profile, 'dentistaprofile', None)
        )

        if (
            user_form.is_valid() and
            profile_form.is_valid() and
            foto_form.is_valid() and
            dentista_form.is_valid()
        ):
            user_form.save()
            profile_form.save()
            foto_form.save()

            dentista = dentista_form.save(commit=False)
            dentista.profile = profile
            dentista.save()

            return redirect('Minha_Conta')

        return render(request, self.template_name, {
            'user_form': user_form,
            'profile_form': profile_form,
            'foto_form': foto_form,
            'dentista_form': dentista_form
        })

class Dados_Da_ClinicaView(LoginRequiredMixin, UserPassesTestMixin, View):
    template_name = 'Clinica_Sorriso/Minha_Conta/Dados_Clinica.html'

    def test_func(self):
        return self.request.user.groups.filter(name='Administrador').exists()

    def get(self, request):
        clinica, created = DadosDaClinica.objects.get_or_create(id=1)
        return render(request, self.template_name, {
            'clinica': clinica
        })

    
class Editar_Dados_Da_ClinicaView(LoginRequiredMixin, UserPassesTestMixin, View):
    template_name = 'Clinica_Sorriso/Minha_Conta/Editar_Dados_Clinica.html'

    def test_func(self):
        return self.request.user.groups.filter(name='Administrador').exists()

    def get(self, request):
        clinica, created = DadosDaClinica.objects.get_or_create(id=1)
        clinica_form = DadosDaClinicaForm(instance=clinica)
        return render(request, self.template_name, {
            'clinica_form': clinica_form
        })

    def post(self, request):
        clinica, created = DadosDaClinica.objects.get_or_create(id=1)
        clinica_form = DadosDaClinicaForm(
            request.POST,
            request.FILES,
            instance=clinica
        )

        if clinica_form.is_valid():
            clinica_form.save()
            return redirect('dados_clinica')

        return render(request, self.template_name, {
            'clinica_form': clinica_form
        })

# Estoque

# Lista todos os produtos do estoque
class ProdutoEstoqueListView(LoginRequiredMixin, ListView):
    model = ProdutoEstoque
    template_name = 'Clinica_Sorriso/Estoque/Estoque.html'
    context_object_name = 'produtos'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Filtra os produtos onde a quantidade atual é menor que a ideal
        produtos_baixos = [
            p for p in self.get_queryset() 
            if p.quantidade_atual < p.quantidade_ideal
        ]
        
        # Adiciona a lista e a contagem ao contexto do HTML
        context['produtos_baixos'] = produtos_baixos
        context['total_criticos'] = len(produtos_baixos)
        return context


def Adicionar_Estoque(request):
    ProdutoFormSet = modelformset_factory(
        ProdutoEstoque,
        form=ProdutoEstoqueForm,
        formset=BaseProdutoEstoqueFormSet,
        extra=0  # 🔥 IMPORTANTE
    )

    if request.method == 'POST':
        data = request.POST.copy()

        total_forms = int(data.get('form-TOTAL_FORMS', 0))
        forms_validos = []

        # 🔥 REMOVE FORMULÁRIOS SEM NOME
        for i in range(total_forms):
            nome = data.get(f'form-{i}-nome', '').strip()
            if nome:
                forms_validos.append(i)

        # ❌ Nenhum produto preenchido
        if not forms_validos:
            messages.error(request, "Adicione pelo menos um produto.")
            return redirect('Estoque')

        # 🔁 REINDEXA
        data['form-TOTAL_FORMS'] = len(forms_validos)
        data['form-INITIAL_FORMS'] = 0

        for novo_i, velho_i in enumerate(forms_validos):
            for campo in ['nome', 'quantidade_atual', 'quantidade_ideal']:
                data[f'form-{novo_i}-{campo}'] = data.get(f'form-{velho_i}-{campo}')

        formset = ProdutoFormSet(data, queryset=ProdutoEstoque.objects.none())

        if formset.is_valid():
            try:
                with transaction.atomic():
                    formset.save()
                messages.success(request, "Produtos adicionados com sucesso!")
                return redirect('Estoque')
            except Exception:
                messages.error(request, "Erro ao salvar no banco de dados.")
        else:
            erros_ja_enviados = set()

            # 🔥 ERROS DO FORMSET (duplicados na lista)
            for error in formset.non_form_errors():
                if error not in erros_ja_enviados:
                    messages.error(request, error)
                    erros_ja_enviados.add(error)

            # 🔥 ERROS DOS FORMS
            for form in formset:
                for error in form.non_field_errors():
                    if error not in erros_ja_enviados:
                        messages.error(request, error)
                        erros_ja_enviados.add(error)

                for field in form:
                    for error in field.errors:
                        if error not in erros_ja_enviados:
                            messages.error(request, error)
                            erros_ja_enviados.add(error)

    else:
        formset = ProdutoFormSet(queryset=ProdutoEstoque.objects.none())

    produtos = ProdutoEstoque.objects.all()
    return render(request, 'Clinica_Sorriso/Estoque/Estoque.html', {
        'formset': formset,
        'produtos': produtos
    })

def Entrada_Produto_Existente(request):
    if request.method == 'POST':
        produto_id = request.POST.get('produto_id')
        quantidade = int(request.POST.get('quantidade'))

        produto = get_object_or_404(ProdutoEstoque, id=produto_id)
        produto.quantidade_atual += quantidade
        produto.save()

        messages.success(
            request,
            f'Entrada realizada: +{quantidade} no produto "{produto.nome}".'
        )

    return redirect('Estoque')

def Confirmar_Entrada_Estoque(request):
    if request.method == 'POST':
        dados_json = request.POST.get('json_dados')
        if not dados_json:
            messages.error(request, "Nenhum dado recebido.")
            return redirect('Estoque')

        try:
            produtos_alterados = json.loads(dados_json)
            with transaction.atomic():
                for item in produtos_alterados:
                    produto = get_object_or_404(ProdutoEstoque, id=item['id'])
                    produto.quantidade_atual += int(item['qtd'])
                    produto.save()
            messages.success(request, "Estoque atualizado com sucesso!")
        except Exception as e:
            messages.error(request, f"Erro ao atualizar: {e}")

    return redirect('Estoque')


# Edita produto existente
@login_required
def Editar_Produto_Estoque(request, pk):
    if not request.user.groups.filter(name='Administrador').exists():
        messages.error(request, "Você não tem permissão para editar produtos.")
        return redirect('Estoque')

    produto = get_object_or_404(ProdutoEstoque, pk=pk)

    if request.method == 'POST':
        nome = request.POST.get('nome')
        quantidade_ideal = request.POST.get('quantidade_ideal')

        # evita nome duplicado
        if ProdutoEstoque.objects.exclude(pk=pk).filter(nome__iexact=nome).exists():
            messages.error(request, "Já existe um produto com esse nome.")
            return redirect('Estoque')

        produto.nome = nome
        produto.quantidade_ideal = quantidade_ideal
        produto.save()

        messages.success(request, "Produto atualizado com sucesso.")
        return redirect('Estoque')

class ProdutoEstoqueDeleteView(LoginRequiredMixin, UserPassesTestMixin, SuccessMessageMixin, DeleteView):
    model = ProdutoEstoque
    success_url = reverse_lazy('Estoque')
    success_message = "Produto excluído com sucesso!" # O MIXIN USA ISSO

    def delete(self, request, *args, **kwargs):
        # Esta é a forma correta de garantir a mensagem em uma DeleteView
        messages.success(self.request, self.success_message)
        return super().delete(request, *args, **kwargs)

    def test_func(self):
        return self.request.user.groups.filter(name='Administrador').exists()


@login_required
def Retirar_Produto(request, pk):
    produto = get_object_or_404(ProdutoEstoque, pk=pk)

    if request.method == 'POST':
        qtd = int(request.POST.get('quantidade_retirada'))

        if qtd > produto.quantidade_atual:
            messages.error(
                request,
                "Estoque insuficiente para a quantidade solicitada.",
                extra_tags='retirada'
            )
            return redirect(f"{reverse('Estoque')}?produto={produto.id}&nome={produto.nome}&quantidade={produto.quantidade_atual}")


        # ✅ SÓ CHEGA AQUI SE FOR VÁLIDO
        RetiradaEstoque.objects.create(
            produto=produto,
            quantidade_retirada=qtd,
            retirado_por=request.user,
            autorizado_por=request.user
        )

        messages.success(
            request,
            "Retirada realizada com sucesso.",
            extra_tags='retirada'
        )
        return redirect('Estoque')

@login_required
def Historico_De_Retiradas(request):
    # .select_related evita várias consultas ao banco de dados
    retiradas = RetiradaEstoque.objects.select_related('produto', 'retirado_por', 'autorizado_por').all().order_by('-data_retirada')
    
    dados = []
    for r in retiradas:
        dados.append({
            'produto': r.produto.nome,
            'quantidade': r.quantidade_retirada,
            # Pegamos o nome completo. Se não tiver, pegamos o username.
            'retirado_nome': r.retirado_por.get_full_name() or r.retirado_por.username if r.retirado_por else "N/A",
            'autorizado_nome': r.autorizado_por.get_full_name() or r.autorizado_por.username if r.autorizado_por else "Admin",
            'data': r.data_retirada.strftime('%d/%m/%Y')
        })
    
    return JsonResponse({'historico': dados})

@login_required
def Exportar_Estoque_Excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Relatório de Estoque"

    # 1. Cabeçalho
    cabecalho = ['PRODUTO', 'QUANTIDADE', 'QTD IDEAL', 'NÍVEL CRÍTICO']
    ws.append(cabecalho)

    # Estilos para o Cabeçalho (Fundo azul, letra branca e negrito)
    estilo_cabecalho = Font(bold=True, color="FFFFFF")
    preenchimento_cabecalho = PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")
    alinhamento_central = Alignment(horizontal="center", vertical="center")

    for cell in ws[1]: # Aplica na primeira linha
        cell.font = estilo_cabecalho
        cell.fill = preenchimento_cabecalho
        cell.alignment = alinhamento_central

    # 2. Definição das fontes para os dados
    fonte_vermelha = Font(color="FF0000", bold=True)
    fonte_verde = Font(color="00B050", bold=True)

    # 3. Inserção dos Dados
    for produto in ProdutoEstoque.objects.all().order_by('nome'):
        if produto.quantidade_atual < produto.quantidade_ideal:
            nivel_critico = "Sim"
            cor_status = fonte_vermelha
        else:
            nivel_critico = "Não"
            cor_status = fonte_verde
        
        ws.append([
            produto.nome.upper(), # Nome em maiúsculo para padronizar
            produto.quantidade_atual, 
            produto.quantidade_ideal, 
            nivel_critico
        ])

        # Aplica a cor na última célula inserida (Coluna 4) e centraliza
        ws.cell(row=ws.max_row, column=4).font = cor_status
        ws.cell(row=ws.max_row, column=2).alignment = alinhamento_central
        ws.cell(row=ws.max_row, column=3).alignment = alinhamento_central
        ws.cell(row=ws.max_row, column=4).alignment = alinhamento_central

    # 4. FORMATAÇÃO: Ajuste automático da largura das colunas
    # Isso evita que o texto fique cortado
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter # Pega a letra da coluna (A, B, C...)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 5)
        ws.column_dimensions[column].width = adjusted_width

    # 5. Configuração do Download
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Estoque Clínica Sorriso & Cia.xlsx"'
    wb.save(response)
    
    return response


def preparar_compromissos_grid(compromissos, horas):
    compromissos_grid = []

    mapa_horas = {}

    for index, hora in enumerate(horas):
        mapa_horas[hora["valor"]] = index + 2

    for compromisso in compromissos:
        if not compromisso.data_inicio:
            continue

        grid_col = compromisso.data_inicio.weekday() + 2

        if compromisso.dia_inteiro:
            hora_inicio_str = "08:00"
            grid_row = mapa_horas.get(hora_inicio_str, 2)
            grid_span = len(horas)
        else:
            if not compromisso.hora_inicio or not compromisso.hora_fim:
                continue

            hora_inicio_str = compromisso.hora_inicio.strftime("%H:%M")
            grid_row = mapa_horas.get(hora_inicio_str)

            if not grid_row:
                continue

            inicio_minutos = compromisso.hora_inicio.hour * 60 + compromisso.hora_inicio.minute
            fim_minutos = compromisso.hora_fim.hour * 60 + compromisso.hora_fim.minute
            duracao_minutos = fim_minutos - inicio_minutos

            if duracao_minutos <= 0:
                duracao_minutos = 15

            grid_span = max(1, duracao_minutos // 15)

        compromisso.grid_col = grid_col
        compromisso.grid_row = grid_row
        compromisso.grid_span = grid_span

        compromissos_grid.append(compromisso)

    return compromissos_grid

@login_required
def agenda(request):
    semana = int(request.GET.get("semana", 0))
    data_str = request.GET.get("data")
    profissional_id = request.GET.get("profissional") or ""

    if data_str:
        data_base = datetime.strptime(data_str, "%Y-%m-%d").date()
    else:
        data_base = date.today() + timedelta(weeks=semana)

    inicio_semana = data_base - timedelta(days=data_base.weekday())
    fim_semana = inicio_semana + timedelta(days=6)

    # ================= PROFISSIONAIS =================
    profissionais = (
        Profissional.objects
        .all()
        .order_by("nome")
    )

    profissional_selecionado = None

    if profissional_id:
        profissional_selecionado = (
            profissionais
            .filter(id=profissional_id)
            .first()
        )

    # ================= ETIQUETAS / RÓTULOS =================
    etiquetas = (
        Etiqueta.objects
        .all()
        .order_by("nome")
    )

    encaixes = (
        EncaixeAgenda.objects
        .select_related("paciente", "profissional", "etiqueta")
        .all()
        .order_by("-urgente", "data_desejada", "-criado_em")
    )

    alertas_retorno = (
        AlertaRetorno.objects
        .select_related("paciente", "profissional")
        .all()
        .order_by("data_retorno", "-criado_em")
    )

    # ================= HORÁRIOS =================
    horas = gerar_horas()

    # ================= DIAS =================
    nomes_dias = ["Seg", "Ter", "Qua", "Qui", "Sex", "Sáb", "Dom"]

    dias = []

    for i in range(7):
        dia = inicio_semana + timedelta(days=i)

        dias.append({
            "data": dia,
            "nome": nomes_dias[dia.weekday()],
            "num": dia.day,
        })

    # ================= AGENDAMENTOS / CONSULTAS =================
    agendamentos = (
        Agendamento.objects
        .select_related("paciente", "profissional")
        .filter(data__range=[inicio_semana, fim_semana])
    )

    if profissional_id:
        agendamentos = agendamentos.filter(profissional_id=profissional_id)

    agendamentos = preparar_agendamentos_grid(agendamentos, horas)

    # ================= COMPROMISSOS =================
    compromissos = (
        CompromissoAgenda.objects
        .select_related("profissional")
        .filter(data_inicio__range=[inicio_semana, fim_semana])
        .order_by("data_inicio", "hora_inicio")
    )

    if profissional_id:
        compromissos = compromissos.filter(profissional_id=profissional_id)

    compromissos = preparar_compromissos_grid(compromissos, horas)

    context = {
        "horas": horas,
        "dias": dias,
        "today": date.today(),

        # consultas
        "agendamentos": agendamentos,

        # compromissos
        "compromissos": compromissos,

        # rótulos / etiquetas
        "etiquetas": etiquetas,

        # semana
        "semana": semana,
        "data_base": data_base,
        "inicio_semana": inicio_semana,
        "fim_semana": fim_semana,

        # expediente
        "inicio_expediente": "08:00",
        "fim_expediente": "18:00",

        # profissionais
        "profissionais": profissionais,
        "profissional_id": profissional_id,
        "profissional_selecionado": profissional_selecionado,

        "encaixes": encaixes,

        "alertas_retorno": alertas_retorno,

    }

    return render(
        request,
        "Clinica_Sorriso/agenda/agenda.html",
        context
    )

@login_required
def criar_agendamento(request):
    profissional_id = request.GET.get("profissional") or request.POST.get("profissional")

    if request.method == "POST":
        tipo_agendamento = request.POST.get("tipo_agendamento", "consulta")

        if tipo_agendamento == "compromisso":
            titulo = request.POST.get("titulo_compromisso", "").strip()
            descricao = request.POST.get("descricao_compromisso", "").strip()
            profissional_compromisso_id = request.POST.get("profissional_compromisso")

            data_inicio = request.POST.get("compromisso_data_inicio")
            hora_inicio = request.POST.get("compromisso_hora_inicio")
            data_fim = request.POST.get("compromisso_data_fim")
            hora_fim = request.POST.get("compromisso_hora_fim")

            dia_inteiro = request.POST.get("dia_inteiro") == "on"
            repetir = request.POST.get("repetir_compromisso") == "on"

            tipo_repeticao = request.POST.get("tipo_repeticao_compromisso") if repetir else None
            terminar_repeticao = request.POST.get("terminar_repeticao_compromisso") or "nunca"
            data_fim_repeticao = request.POST.get("data_fim_repeticao_compromisso") or None

            disponibilidade = request.POST.get("disponibilidade_compromisso") or "ocupado"
            privacidade = request.POST.get("privacidade_compromisso") or "privado"
            alerta = request.POST.get("alerta_compromisso") or "sem_alerta"

            alerta_quantidade = request.POST.get("alerta_quantidade") or None
            alerta_unidade = request.POST.get("alerta_unidade") or None

            if not titulo:
                return JsonResponse({
                    "ok": False,
                    "erro": "Informe o título do compromisso."
                }, status=400)

            if not profissional_compromisso_id:
                return JsonResponse({
                    "ok": False,
                    "erro": "Selecione a agenda/profissional."
                }, status=400)

            try:
                compromisso = CompromissoAgenda.objects.create(
                    titulo=titulo,
                    descricao=descricao,
                    profissional_id=profissional_compromisso_id,
                    data_inicio=data_inicio,
                    hora_inicio=None if dia_inteiro else hora_inicio,
                    data_fim=data_fim,
                    hora_fim=None if dia_inteiro else hora_fim,
                    dia_inteiro=dia_inteiro,
                    repetir=repetir,
                    tipo_repeticao=tipo_repeticao,
                    terminar_repeticao=terminar_repeticao,
                    data_fim_repeticao=data_fim_repeticao,
                    disponibilidade=disponibilidade,
                    privacidade=privacidade,
                    alerta=alerta,
                    alerta_quantidade=alerta_quantidade,
                    alerta_unidade=alerta_unidade,
                )
            except Exception as e:
                return JsonResponse({
                    "ok": False,
                    "erro": str(e)
                }, status=400)

            if request.headers.get("x-requested-with") == "XMLHttpRequest":
                return JsonResponse({
                    "ok": True,
                    "mensagem": "Compromisso criado com sucesso.",
                    "tipo": "compromisso",
                    "id": compromisso.id,
                })

            return redirect("agenda")

        form = AgendamentoForm(request.POST)

        if form.is_valid():
            nome_paciente = form.cleaned_data["paciente_nome"].strip()

            paciente = Paciente.objects.filter(nome__iexact=nome_paciente).first()

            if not paciente:
                paciente = Paciente.objects.create(nome=nome_paciente)

            agendamento = form.save(commit=False)
            agendamento.paciente = paciente
            agendamento.save()

            # Marca o alerta de retorno como agendado, se esse agendamento veio de um alerta
            alerta_retorno_id = request.POST.get("alerta_retorno_id")

            if alerta_retorno_id:
                alerta_retorno = AlertaRetorno.objects.filter(id=alerta_retorno_id).first()

                if alerta_retorno:
                    alerta_retorno.status = "agendado"
                    alerta_retorno.save(update_fields=["status"])

            if request.headers.get("x-requested-with") == "XMLHttpRequest":
                return JsonResponse({
                    "ok": True,
                    "mensagem": "Agendamento criado com sucesso.",
                    "tipo": "consulta",
                    "id": agendamento.id,
                })

            return redirect("agenda")

        else:
            print("FORM INVALIDO:", form.errors)

            if request.headers.get("x-requested-with") == "XMLHttpRequest":
                return JsonResponse({
                    "ok": False,
                    "erros": form.errors
                }, status=400)

    else:
        initial = {
            "data": request.GET.get("dia"),
            "hora_inicio": request.GET.get("hora"),
        }

        if profissional_id:
            initial["profissional"] = profissional_id

        form = AgendamentoForm(initial=initial)

    return render(request, "Clinica_Sorriso/agenda/criar_agendamento.html", {
        "form": form,
        "profissional_id": profissional_id,
    })


from django.http import JsonResponse
from django.views.decorators.http import require_POST

@require_POST
def alterar_status_agendamento(request):
    id = request.POST.get("id")
    status = request.POST.get("status")

    try:
        ag = Agendamento.objects.get(id=id)
        ag.status = status
        ag.save()

        return JsonResponse({
            "success": True,
            "novo_status": ag.get_status_display()
        })

    except Agendamento.DoesNotExist:
        return JsonResponse({"success": False})


@login_required
@require_POST
def salvar_retorno(request):
    try:
        paciente_id = request.POST.get("paciente_id")
        paciente_nome = request.POST.get("paciente_nome", "").strip()

        profissional_id = request.POST.get("profissional", "").strip()
        tipo_retorno = request.POST.get("tipo_retorno", "data_especifica").strip()
        data_retorno = request.POST.get("data_retorno", "").strip()
        motivo = request.POST.get("motivo", "").strip()

        if not paciente_id and not paciente_nome:
            return JsonResponse({
                "ok": False,
                "erro": "Informe o paciente."
            }, status=400)

        if not profissional_id:
            return JsonResponse({
                "ok": False,
                "erro": "Selecione o profissional."
            }, status=400)

        if not tipo_retorno:
            return JsonResponse({
                "ok": False,
                "erro": "Selecione quando o paciente deve retornar."
            }, status=400)

        paciente = None

        if paciente_id:
            paciente = Paciente.objects.filter(id=paciente_id).first()

        if not paciente and paciente_nome:
            paciente = Paciente.objects.filter(nome__iexact=paciente_nome).first()

        if not paciente:
            return JsonResponse({
                "ok": False,
                "erro": "Paciente não encontrado. Cadastre o paciente antes de salvar o retorno."
            }, status=404)

        profissional = Profissional.objects.filter(id=profissional_id).first()

        if not profissional:
            return JsonResponse({
                "ok": False,
                "erro": "Profissional não encontrado."
            }, status=404)

        data_obj = None

        if tipo_retorno == "data_especifica":
            if not data_retorno:
                return JsonResponse({
                    "ok": False,
                    "erro": "Informe a data aproximada para retorno."
                }, status=400)

            data_obj = datetime.strptime(data_retorno, "%Y-%m-%d").date()

        else:
            hoje = date.today()

            if tipo_retorno == "7_dias":
                data_obj = hoje + timedelta(days=7)
            elif tipo_retorno == "15_dias":
                data_obj = hoje + timedelta(days=15)
            elif tipo_retorno == "30_dias":
                data_obj = hoje + timedelta(days=30)
            elif tipo_retorno == "3_meses":
                data_obj = hoje + timedelta(days=90)
            elif tipo_retorno == "6_meses":
                data_obj = hoje + timedelta(days=180)

        alerta = AlertaRetorno.objects.create(
            paciente=paciente,
            profissional=profissional,
            tipo_retorno=tipo_retorno,
            data_retorno=data_obj,
            motivo=motivo,
            status="pendente"
        )

        return JsonResponse({
            "ok": True,
            "mensagem": "Alerta de retorno salvo com sucesso.",
            "alerta_id": alerta.id
        })

    except Exception as e:
        return JsonResponse({
            "ok": False,
            "erro": str(e)
        }, status=500)
    
@login_required
def agenda_parcial(request):
    semana = int(request.GET.get("semana", 0))
    data_str = request.GET.get("data")
    profissional_id = request.GET.get("profissional") or ""

    if data_str:
        data_base = datetime.strptime(data_str, "%Y-%m-%d").date()
    else:
        data_base = date.today() + timedelta(weeks=semana)

    inicio_semana = data_base - timedelta(days=data_base.weekday())
    fim_semana = inicio_semana + timedelta(days=6)

    # ================= PROFISSIONAIS =================
    profissionais = (
        Profissional.objects
        .all()
        .order_by("nome")
    )

    profissional_selecionado = None

    if profissional_id:
        profissional_selecionado = (
            profissionais
            .filter(id=profissional_id)
            .first()
        )

    # ================= HORÁRIOS =================
    horas = gerar_horas()

    # ================= DIAS =================
    nomes_dias = ["Seg", "Ter", "Qua", "Qui", "Sex", "Sáb", "Dom"]

    dias = []

    for i in range(7):
        dia = inicio_semana + timedelta(days=i)

        dias.append({
            "data": dia,
            "nome": nomes_dias[dia.weekday()],
            "num": dia.day,
        })

    # ================= AGENDAMENTOS =================
    agendamentos = (
        Agendamento.objects
        .select_related("paciente", "profissional")
        .filter(data__range=[inicio_semana, fim_semana])
    )

    if profissional_id:
        agendamentos = agendamentos.filter(profissional_id=profissional_id)

    agendamentos = preparar_agendamentos_grid(agendamentos, horas)

    return render(
        request,
        "Clinica_Sorriso/agenda/agenda_parcial.html",
        {
            "horas": horas,
            "dias": dias,
            "today": date.today(),
            "agendamentos": agendamentos,
            "semana": semana,
            "data_base": data_base,
            "inicio_expediente": "08:00",
            "fim_expediente": "18:00",

            # profissionais
            "profissionais": profissionais,
            "profissional_id": profissional_id,
            "profissional_selecionado": profissional_selecionado,
        }
    )


def gerar_horas():
    horas = []

    for h in range(8, 19):

        horas.append({
            "valor": f"{h:02}:00",
            "cheia": True,
            "fechado": False,
        })

        if h < 18:
            for minuto in ("15", "30", "45"):
                horas.append({
                    "valor": f"{h:02}:{minuto}",
                    "cheia": False,
                    "fechado": False,
                })

    return horas


import math
from datetime import date, timedelta, datetime


def preparar_agendamentos_grid(agendamentos, horas):
    mapa_horas = {
        h["valor"]: index + 1
        for index, h in enumerate(horas)
    }

    for ag in agendamentos:
        hora_inicio = ag.hora_inicio.strftime("%H:%M")
        hora_fim = ag.hora_fim.strftime("%H:%M") if ag.hora_fim else None

        ag.grid_row = mapa_horas.get(hora_inicio, 1) + 1

        # coluna 1 é a coluna das horas, então os dias começam na coluna 2
        ag.grid_col = ag.data.weekday() + 2

        if ag.hora_fim:
            inicio_dt = datetime.combine(ag.data, ag.hora_inicio)
            fim_dt = datetime.combine(ag.data, ag.hora_fim)

            minutos = (fim_dt - inicio_dt).total_seconds() / 60

            if minutos <= 0:
                minutos = 15
        else:
            minutos = 15

        ag.grid_span = max(1, math.ceil(minutos / 15))

        # ================= DADOS EXTRAS PARA O MODAL =================
        if ag.paciente:
            ag.paciente_nome = ag.paciente.nome

            # tenta pegar celular, telefone ou whatsapp, conforme existir no seu model
            ag.paciente_celular = (
                getattr(ag.paciente, "celular", "") or
                getattr(ag.paciente, "telefone", "") or
                getattr(ag.paciente, "whatsapp", "")
            )
        else:
            ag.paciente_nome = ""
            ag.paciente_celular = ""

    return agendamentos


from django.shortcuts import get_object_or_404
@login_required
def detalhes_agendamento(request, id):
    ag = (
        Agendamento.objects
        .select_related("paciente", "profissional", "etiqueta")
        .filter(id=id)
        .first()
    )

    if not ag:
        return HttpResponse("Agendamento não encontrado.", status=404)

    if ag.paciente:
        ag.paciente_nome = getattr(ag.paciente, "nome", str(ag.paciente))

        ag.paciente_celular = (
            getattr(ag.paciente, "celular", "") or
            getattr(ag.paciente, "telefone", "") or
            getattr(ag.paciente, "whatsapp", "")
        )
    else:
        ag.paciente_nome = ""
        ag.paciente_celular = ""

    etiquetas = Etiqueta.objects.all().order_by("nome")

    return render(
        request,
        "Clinica_Sorriso/agenda/detalhes_agendamento.html",
        {
            "ag": ag,
            "etiquetas": etiquetas,
        }
    )


from django.http import JsonResponse, HttpResponse

@login_required
def buscar_pacientes_agenda(request):
    termo = request.GET.get("q", "").strip()

    pacientes = []

    if termo:
        pacientes_qs = (
            Paciente.objects
            .filter(nome__icontains=termo)
            .order_by("nome")[:10]
        )

        pacientes = [
            {
                "id": paciente.id,
                "nome": paciente.nome,
                "cpf": getattr(paciente, "cpf", "") or "",
                "celular": getattr(paciente, "celular", "") or "",
            }
            for paciente in pacientes_qs
        ]

    return JsonResponse({
        "pacientes": pacientes
    })

@login_required
def criar_paciente(request):
    origem = request.GET.get("origem") or request.POST.get("origem")

    if request.method == "POST":
        form = PacienteForm(request.POST, request.FILES)

        if form.is_valid():
            form.save()

            if origem == "agenda":
                return redirect("agenda")

            return redirect("agenda")

        else:
            print("FORM PACIENTE INVALIDO:", form.errors)

    else:
        nome_inicial = request.GET.get("nome", "").strip()

        form = PacienteForm(
            initial={
                "nome": nome_inicial
            }
        )

    return render(
        request,
        "Clinica_Sorriso/pacientes/criar_paciente.html",
        {
            "form": form,
            "origem": origem,
        }
    )

from datetime import datetime, timedelta
from django.http import JsonResponse

@login_required
def horarios_livres_agenda(request):
    data_str = request.GET.get("data")
    profissional_id = request.GET.get("profissional")
    duracao = int(request.GET.get("duracao") or 30)

    if not data_str or not profissional_id:
        return JsonResponse({
            "ok": False,
            "erro": "Data e profissional são obrigatórios."
        }, status=400)

    data_consulta = datetime.strptime(data_str, "%Y-%m-%d").date()

    inicio_expediente = datetime.combine(
        data_consulta,
        datetime.strptime("08:00", "%H:%M").time()
    )

    fim_expediente = datetime.combine(
        data_consulta,
        datetime.strptime("18:00", "%H:%M").time()
    )

    agendamentos = (
        Agendamento.objects
        .filter(
            data=data_consulta,
            profissional_id=profissional_id
        )
        .order_by("hora_inicio")
    )

    ocupados = []

    for ag in agendamentos:
        inicio = datetime.combine(data_consulta, ag.hora_inicio)
        fim = datetime.combine(data_consulta, ag.hora_fim)

        ocupados.append((inicio, fim))

    horarios_livres = []
    atual = inicio_expediente

    while atual + timedelta(minutes=duracao) <= fim_expediente:
        fim_teste = atual + timedelta(minutes=duracao)

        conflito = False

        for inicio_ocupado, fim_ocupado in ocupados:
            if atual < fim_ocupado and fim_teste > inicio_ocupado:
                conflito = True
                break

        if not conflito:
            horarios_livres.append({
                "inicio": atual.strftime("%H:%M"),
                "fim": fim_teste.strftime("%H:%M"),
                "texto": f"{atual.strftime('%H:%M')} às {fim_teste.strftime('%H:%M')}"
            })

        atual += timedelta(minutes=15)

    return JsonResponse({
        "ok": True,
        "horarios": horarios_livres
    })

from django.http import JsonResponse
from django.views.decorators.http import require_POST
from .models import Etiqueta

@login_required
@require_POST
def criar_etiqueta_agenda(request):
    nome = request.POST.get("nome", "").strip()
    cor = request.POST.get("cor", "#28a745").strip()

    if not nome:
        return JsonResponse({
            "ok": False,
            "erro": "Informe o nome do rótulo."
        }, status=400)

    etiqueta, criada = Etiqueta.objects.get_or_create(
        nome=nome,
        defaults={
            "cor": cor or "#28a745"
        }
    )

    if not criada:
        etiqueta.cor = cor or etiqueta.cor
        etiqueta.save()

    return JsonResponse({
        "ok": True,
        "etiqueta": {
            "id": etiqueta.id,
            "nome": etiqueta.nome,
            "cor": etiqueta.cor,
        }
    })

from django.views.decorators.http import require_POST
from django.http import JsonResponse

@login_required
@require_POST
def editar_compromisso_agenda(request, compromisso_id):
    compromisso = CompromissoAgenda.objects.filter(id=compromisso_id).first()

    if not compromisso:
        return JsonResponse({
            "ok": False,
            "erro": "Compromisso não encontrado."
        }, status=404)

    titulo = request.POST.get("titulo", "").strip()
    descricao = request.POST.get("descricao", "").strip()
    profissional_id = request.POST.get("profissional")

    data_inicio = request.POST.get("data_inicio")
    data_fim = request.POST.get("data_fim")
    hora_inicio = request.POST.get("hora_inicio")
    hora_fim = request.POST.get("hora_fim")

    dia_inteiro = request.POST.get("dia_inteiro") == "on"
    repetir = request.POST.get("repetir") == "on"

    tipo_repeticao = request.POST.get("tipo_repeticao") if repetir else None
    terminar_repeticao = request.POST.get("terminar_repeticao") or "nunca"
    data_fim_repeticao = request.POST.get("data_fim_repeticao") or None

    disponibilidade = request.POST.get("disponibilidade") or "ocupado"
    privacidade = request.POST.get("privacidade") or "privado"
    alerta = request.POST.get("alerta") or "sem_alerta"
    alerta_quantidade = request.POST.get("alerta_quantidade") or None
    alerta_unidade = request.POST.get("alerta_unidade") or None

    if not titulo:
        return JsonResponse({
            "ok": False,
            "erro": "Informe o título do compromisso."
        }, status=400)

    if not profissional_id:
        return JsonResponse({
            "ok": False,
            "erro": "Selecione a agenda/profissional."
        }, status=400)

    if not data_inicio or not data_fim:
        return JsonResponse({
            "ok": False,
            "erro": "Informe a data de início e fim."
        }, status=400)

    if not dia_inteiro and (not hora_inicio or not hora_fim):
        return JsonResponse({
            "ok": False,
            "erro": "Informe o horário de início e fim."
        }, status=400)

    try:
        compromisso.titulo = titulo
        compromisso.descricao = descricao
        compromisso.profissional_id = profissional_id
        compromisso.data_inicio = data_inicio
        compromisso.data_fim = data_fim

        compromisso.dia_inteiro = dia_inteiro

        if dia_inteiro:
            compromisso.hora_inicio = None
            compromisso.hora_fim = None
        else:
            compromisso.hora_inicio = hora_inicio
            compromisso.hora_fim = hora_fim

        compromisso.repetir = repetir
        compromisso.tipo_repeticao = tipo_repeticao
        compromisso.terminar_repeticao = terminar_repeticao
        compromisso.data_fim_repeticao = data_fim_repeticao

        compromisso.disponibilidade = disponibilidade
        compromisso.privacidade = privacidade
        compromisso.alerta = alerta
        compromisso.alerta_quantidade = alerta_quantidade
        compromisso.alerta_unidade = alerta_unidade

        compromisso.save()

    except Exception as e:
        return JsonResponse({
            "ok": False,
            "erro": str(e)
        }, status=400)

    return JsonResponse({
        "ok": True,
        "mensagem": "Compromisso atualizado com sucesso."
    })


@login_required
@require_POST
def atualizar_rotulo_agendamento(request, agendamento_id):
    agendamento = Agendamento.objects.filter(id=agendamento_id).first()

    if not agendamento:
        return JsonResponse({
            "ok": False,
            "erro": "Agendamento não encontrado."
        }, status=404)

    etiqueta_id = request.POST.get("etiqueta_id")

    if etiqueta_id:
        etiqueta = Etiqueta.objects.filter(id=etiqueta_id).first()

        if not etiqueta:
            return JsonResponse({
                "ok": False,
                "erro": "Rótulo não encontrado."
            }, status=404)

        agendamento.etiqueta = etiqueta
    else:
        agendamento.etiqueta = None

    agendamento.save(update_fields=["etiqueta"])

    return JsonResponse({
        "ok": True,
        "etiqueta_id": etiqueta_id or ""
    })

from datetime import datetime, timedelta
from django.http import JsonResponse
from django.views.decorators.http import require_POST
from django.contrib.auth.decorators import login_required
from .models import Agendamento, Paciente, Profissional
from django.core.exceptions import ValidationError

@login_required
@require_POST
def duplicar_consulta_agenda(request):
    try:
        paciente_id = request.POST.get("paciente_id")
        paciente_nome = request.POST.get("paciente_nome", "").strip()
        profissional_id = request.POST.get("profissional")

        data = request.POST.get("data")
        hora_inicio = request.POST.get("hora_inicio")
        duracao = request.POST.get("duracao") or 30
        observacoes = request.POST.get("observacoes", "").strip()

        if not paciente_nome:
            return JsonResponse({"ok": False, "erro": "Informe o paciente."}, status=400)

        if not profissional_id:
            return JsonResponse({"ok": False, "erro": "Selecione o profissional."}, status=400)

        if not data:
            return JsonResponse({"ok": False, "erro": "Informe a data da consulta."}, status=400)

        if not hora_inicio:
            return JsonResponse({"ok": False, "erro": "Informe a hora de início."}, status=400)

        paciente = None

        if paciente_id:
            paciente = Paciente.objects.filter(id=paciente_id).first()

        if not paciente:
            paciente = Paciente.objects.filter(nome__iexact=paciente_nome).first()

        if not paciente:
            return JsonResponse({"ok": False, "erro": "Paciente não encontrado."}, status=404)

        profissional = Profissional.objects.filter(id=profissional_id).first()

        if not profissional:
            return JsonResponse({"ok": False, "erro": "Profissional não encontrado."}, status=404)

        try:
            duracao = int(duracao)
        except ValueError:
            duracao = 30

        data_obj = datetime.strptime(data, "%Y-%m-%d").date()
        hora_inicio_obj = datetime.strptime(hora_inicio, "%H:%M").time()

        inicio_dt = datetime.combine(data_obj, hora_inicio_obj)
        fim_dt = inicio_dt + timedelta(minutes=duracao)
        hora_fim_obj = fim_dt.time()

        agendamento = Agendamento(
            paciente=paciente,
            profissional=profissional,
            data=data_obj,
            hora_inicio=hora_inicio_obj,
            hora_fim=hora_fim_obj,
            observacoes=observacoes,
            status="pendente",
        )

        agendamento.full_clean()
        agendamento.save()

        return JsonResponse({
            "ok": True,
            "mensagem": "Consulta duplicada com sucesso.",
            "agendamento_id": agendamento.id,
        })

    except ValidationError as e:
        mensagens = []

        if hasattr(e, "message_dict"):
            for lista_erros in e.message_dict.values():
                mensagens.extend(lista_erros)
        else:
            mensagens = e.messages

        return JsonResponse({
            "ok": False,
            "erro": mensagens[0] if mensagens else "Não foi possível duplicar a consulta."
        }, status=400)

    except Exception as e:
        return JsonResponse({
            "ok": False,
            "erro": str(e),
        }, status=500)


@login_required
@require_POST
def editar_consulta_agenda(request, agendamento_id):
    try:
        agendamento = Agendamento.objects.filter(id=agendamento_id).first()

        if not agendamento:
            return JsonResponse({
                "ok": False,
                "erro": "Consulta não encontrada."
            }, status=404)

        paciente_id = request.POST.get("paciente_id")
        paciente_nome = request.POST.get("paciente_nome", "").strip()
        profissional_id = request.POST.get("profissional")

        data = request.POST.get("data")
        hora_inicio = request.POST.get("hora_inicio")
        duracao = request.POST.get("duracao") or 30
        observacoes = request.POST.get("observacoes", "").strip()
        status = request.POST.get("status", "").strip()
        etiqueta_id = request.POST.get("etiqueta_id", "").strip()

        if not paciente_nome:
            return JsonResponse({
                "ok": False,
                "erro": "Informe o paciente."
            }, status=400)

        if not profissional_id:
            return JsonResponse({
                "ok": False,
                "erro": "Selecione o profissional."
            }, status=400)

        if not data:
            return JsonResponse({
                "ok": False,
                "erro": "Informe a data da consulta."
            }, status=400)

        if not hora_inicio:
            return JsonResponse({
                "ok": False,
                "erro": "Informe a hora de início."
            }, status=400)

        paciente = None

        if paciente_id:
            paciente = Paciente.objects.filter(id=paciente_id).first()

        if paciente:
            # Se o nome foi alterado no modal, atualiza o cadastro do paciente
            if paciente_nome and paciente.nome != paciente_nome:
                paciente.nome = paciente_nome
                paciente.save(update_fields=["nome"])
        else:
            paciente = Paciente.objects.filter(nome__iexact=paciente_nome).first()

        if not paciente:
            return JsonResponse({
                "ok": False,
                "erro": "Paciente não encontrado."
            }, status=404)

        profissional = Profissional.objects.filter(id=profissional_id).first()

        if not profissional:
            return JsonResponse({
                "ok": False,
                "erro": "Profissional não encontrado."
            }, status=404)

        try:
            duracao = int(duracao)
        except ValueError:
            duracao = 30

        data_obj = datetime.strptime(data, "%Y-%m-%d").date()
        hora_inicio_obj = datetime.strptime(hora_inicio, "%H:%M").time()

        inicio_dt = datetime.combine(data_obj, hora_inicio_obj)
        fim_dt = inicio_dt + timedelta(minutes=duracao)
        hora_fim_obj = fim_dt.time()

        agendamento.paciente = paciente
        agendamento.profissional = profissional
        agendamento.data = data_obj
        agendamento.hora_inicio = hora_inicio_obj
        agendamento.hora_fim = hora_fim_obj

        campos_model = [campo.name for campo in Agendamento._meta.fields]

        if "observacoes" in campos_model:
            agendamento.observacoes = observacoes
        elif "observacao" in campos_model:
            agendamento.observacao = observacoes

        if "status" in campos_model and status:
            agendamento.status = status

        if "etiqueta" in campos_model:
            if etiqueta_id:
                etiqueta = Etiqueta.objects.filter(id=etiqueta_id).first()

                if not etiqueta:
                    return JsonResponse({
                        "ok": False,
                        "erro": "Rótulo não encontrado."
                    }, status=404)

                agendamento.etiqueta = etiqueta
            else:
                agendamento.etiqueta = None

        agendamento.full_clean()
        agendamento.save()

        return JsonResponse({
            "ok": True,
            "mensagem": "Consulta salva com sucesso.",
            "agendamento_id": agendamento.id,
        })

    except ValidationError as e:
        mensagens = []

        if hasattr(e, "message_dict"):
            for lista_erros in e.message_dict.values():
                mensagens.extend(lista_erros)
        else:
            mensagens = e.messages

        return JsonResponse({
            "ok": True,
            "mensagem": "Consulta salva com sucesso.",
            "agendamento_id": agendamento.id,
            "salvo": {
                "paciente": str(agendamento.paciente),
                "profissional": str(agendamento.profissional),
                "data": agendamento.data.strftime("%Y-%m-%d"),
                "hora_inicio": agendamento.hora_inicio.strftime("%H:%M"),
                "hora_fim": agendamento.hora_fim.strftime("%H:%M"),
                "status": getattr(agendamento, "status", ""),
                "etiqueta": str(agendamento.etiqueta) if getattr(agendamento, "etiqueta", None) else "",
            }
        })

    except Exception as e:
        return JsonResponse({
            "ok": False,
            "erro": str(e),
        }, status=500)


@login_required
@require_POST
def excluir_consulta_agenda(request, agendamento_id):
    agendamento = Agendamento.objects.filter(id=agendamento_id).first()

    if not agendamento:
        return JsonResponse({
            "ok": False,
            "erro": "Consulta não encontrada."
        }, status=404)

    agendamento.delete()

    return JsonResponse({
        "ok": True,
        "mensagem": "Consulta excluída com sucesso."
    })


@login_required
@require_POST
def salvar_encaixe_agenda(request):
    try:
        paciente_id = request.POST.get("paciente_id")
        paciente_nome = request.POST.get("paciente_nome", "").strip()

        profissional_id = request.POST.get("profissional", "").strip()
        data_desejada = request.POST.get("data_desejada", "").strip()
        qualquer_data = request.POST.get("qualquer_data") == "1"
        turno = request.POST.get("turno", "").strip()
        plano = request.POST.get("plano", "").strip() or "Particular"
        observacao = request.POST.get("observacao", "").strip()
        urgente = request.POST.get("urgente") == "1"
        etiqueta_id = request.POST.get("etiqueta_id", "").strip()

        if not paciente_id and not paciente_nome:
            return JsonResponse({
                "ok": False,
                "erro": "Informe o paciente."
            }, status=400)

        paciente = None

        if paciente_id:
            paciente = Paciente.objects.filter(id=paciente_id).first()

        if not paciente and paciente_nome:
            paciente = Paciente.objects.filter(nome__iexact=paciente_nome).first()

        if not paciente:
            return JsonResponse({
                "ok": False,
                "erro": "Paciente não encontrado. Cadastre o paciente antes de salvar o encaixe."
            }, status=404)

        profissional = None

        if profissional_id:
            profissional = Profissional.objects.filter(id=profissional_id).first()

            if not profissional:
                return JsonResponse({
                    "ok": False,
                    "erro": "Profissional não encontrado."
                }, status=404)

        data_obj = None

        if not qualquer_data:
            if not data_desejada:
                return JsonResponse({
                    "ok": False,
                    "erro": "Informe a data do encaixe ou marque Qualquer data."
                }, status=400)

            data_obj = datetime.strptime(data_desejada, "%Y-%m-%d").date()

        etiqueta = None

        if etiqueta_id:
            etiqueta = Etiqueta.objects.filter(id=etiqueta_id).first()

            if not etiqueta:
                return JsonResponse({
                    "ok": False,
                    "erro": "Rótulo não encontrado."
                }, status=404)

        encaixe = EncaixeAgenda.objects.create(
            paciente=paciente,
            profissional=profissional,
            data_desejada=data_obj,
            qualquer_data=qualquer_data,
            turno=turno,
            plano=plano,
            observacao=observacao,
            urgente=urgente,
            etiqueta=etiqueta,
        )

        return JsonResponse({
            "ok": True,
            "mensagem": "Encaixe salvo com sucesso.",
            "encaixe_id": encaixe.id
        })

    except Exception as e:
        return JsonResponse({
            "ok": False,
            "erro": str(e)
        }, status=500)
    

@login_required
def dados_encaixe_agenda(request, encaixe_id):
    encaixe = (
        EncaixeAgenda.objects
        .select_related("paciente", "profissional", "etiqueta")
        .filter(id=encaixe_id)
        .first()
    )

    if not encaixe:
        return JsonResponse({
            "ok": False,
            "erro": "Encaixe não encontrado."
        }, status=404)

    celular = (
        getattr(encaixe.paciente, "celular", "") or
        getattr(encaixe.paciente, "telefone", "") or
        getattr(encaixe.paciente, "whatsapp", "")
    )

    return JsonResponse({
        "ok": True,
        "encaixe": {
            "id": encaixe.id,
            "paciente_id": encaixe.paciente.id,
            "paciente_nome": encaixe.paciente.nome,
            "paciente_celular": celular,

            "profissional_id": encaixe.profissional.id if encaixe.profissional else "",
            "profissional_nome": encaixe.profissional.nome if encaixe.profissional else "",

            "data": encaixe.data_desejada.strftime("%Y-%m-%d") if encaixe.data_desejada else "",
            "turno": encaixe.turno or "",

            "plano": encaixe.plano or "Particular",
            "observacao": encaixe.observacao or "",

            "etiqueta_id": encaixe.etiqueta.id if encaixe.etiqueta else "",
            "etiqueta_nome": encaixe.etiqueta.nome if encaixe.etiqueta else "",
            "etiqueta_cor": encaixe.etiqueta.cor if encaixe.etiqueta else "",
        }
    })

@login_required
@require_POST
def editar_encaixe_agenda(request, encaixe_id):
    try:
        encaixe = EncaixeAgenda.objects.filter(id=encaixe_id).first()

        if not encaixe:
            return JsonResponse({
                "ok": False,
                "erro": "Encaixe não encontrado."
            }, status=404)

        paciente_id = request.POST.get("paciente_id")
        paciente_nome = request.POST.get("paciente_nome", "").strip()

        profissional_id = request.POST.get("profissional", "").strip()
        data_desejada = request.POST.get("data_desejada", "").strip()
        qualquer_data = request.POST.get("qualquer_data") == "1"
        turno = request.POST.get("turno", "").strip()
        plano = request.POST.get("plano", "").strip() or "Particular"
        observacao = request.POST.get("observacao", "").strip()
        urgente = request.POST.get("urgente") == "1"
        etiqueta_id = request.POST.get("etiqueta_id", "").strip()

        if not paciente_id and not paciente_nome:
            return JsonResponse({
                "ok": False,
                "erro": "Informe o paciente."
            }, status=400)

        paciente = None

        if paciente_id:
            paciente = Paciente.objects.filter(id=paciente_id).first()

        if not paciente and paciente_nome:
            paciente = Paciente.objects.filter(nome__iexact=paciente_nome).first()

        if not paciente:
            return JsonResponse({
                "ok": False,
                "erro": "Paciente não encontrado."
            }, status=404)

        profissional = None

        if profissional_id:
            profissional = Profissional.objects.filter(id=profissional_id).first()

            if not profissional:
                return JsonResponse({
                    "ok": False,
                    "erro": "Profissional não encontrado."
                }, status=404)

        data_obj = None

        if not qualquer_data:
            if not data_desejada:
                return JsonResponse({
                    "ok": False,
                    "erro": "Informe a data do encaixe ou marque Qualquer data."
                }, status=400)

            data_obj = datetime.strptime(data_desejada, "%Y-%m-%d").date()

        etiqueta = None

        if etiqueta_id:
            etiqueta = Etiqueta.objects.filter(id=etiqueta_id).first()

            if not etiqueta:
                return JsonResponse({
                    "ok": False,
                    "erro": "Rótulo não encontrado."
                }, status=404)

        encaixe.paciente = paciente
        encaixe.profissional = profissional
        encaixe.data_desejada = data_obj
        encaixe.qualquer_data = qualquer_data
        encaixe.turno = turno
        encaixe.plano = plano
        encaixe.observacao = observacao
        encaixe.urgente = urgente
        encaixe.etiqueta = etiqueta
        encaixe.save()

        return JsonResponse({
            "ok": True,
            "mensagem": "Encaixe atualizado com sucesso.",
            "encaixe_id": encaixe.id
        })

    except Exception as e:
        return JsonResponse({
            "ok": False,
            "erro": str(e)
        }, status=500)
    

@login_required
@require_POST
def excluir_encaixe_agenda(request, encaixe_id):
    encaixe = EncaixeAgenda.objects.filter(id=encaixe_id).first()

    if not encaixe:
        return JsonResponse({
            "ok": False,
            "erro": "Encaixe não encontrado."
        }, status=404)

    encaixe.delete()

    return JsonResponse({
        "ok": True,
        "mensagem": "Encaixe excluído com sucesso."
    })


@login_required
@require_POST
def excluir_alerta_retorno(request, alerta_id):
    alerta = AlertaRetorno.objects.filter(id=alerta_id).first()

    if not alerta:
        return JsonResponse({
            "ok": False,
            "erro": "Alerta de retorno não encontrado."
        }, status=404)

    alerta.delete()

    return JsonResponse({
        "ok": True,
        "mensagem": "Alerta de retorno excluído com sucesso."
    })

@login_required
def dados_alerta_retorno(request, alerta_id):
    alerta = (
        AlertaRetorno.objects
        .select_related("paciente", "profissional")
        .filter(id=alerta_id)
        .first()
    )

    if not alerta:
        return JsonResponse({
            "ok": False,
            "erro": "Alerta de retorno não encontrado."
        }, status=404)

    celular = (
        getattr(alerta.paciente, "celular", "") or
        getattr(alerta.paciente, "telefone", "") or
        getattr(alerta.paciente, "whatsapp", "")
    )

    return JsonResponse({
        "ok": True,
        "alerta": {
            "id": alerta.id,
            "paciente_id": alerta.paciente.id,
            "paciente_nome": alerta.paciente.nome,
            "paciente_celular": celular,

            "profissional_id": alerta.profissional.id if alerta.profissional else "",
            "profissional_nome": alerta.profissional.nome if alerta.profissional else "",

            "tipo_retorno": alerta.tipo_retorno or "",
            "tipo_retorno_texto": alerta.get_tipo_retorno_display(),

            "data_retorno": alerta.data_retorno.strftime("%Y-%m-%d") if alerta.data_retorno else "",
            "motivo": alerta.motivo or "",
            "status": alerta.status,
        }
    })


@login_required
def exportar_retornos(request):
    from django.http import HttpResponse
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.units import cm

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="alertas_de_retorno.pdf"'

    doc = SimpleDocTemplate(
        response,
        pagesize=A4,
        rightMargin=1.5 * cm,
        leftMargin=1.5 * cm,
        topMargin=1.5 * cm,
        bottomMargin=1.5 * cm,
    )

    elementos = []
    estilos = getSampleStyleSheet()

    titulo = Paragraph("Alertas de Retorno", estilos["Title"])
    elementos.append(titulo)
    elementos.append(Spacer(1, 12))

    alertas = (
        AlertaRetorno.objects
        .select_related("paciente", "profissional")
        .all()
        .order_by("data_retorno", "-criado_em")
    )

    dados = [
        ["Data", "Paciente", "Profissional", "Retornar em", "Status", "Motivo"]
    ]

    for alerta in alertas:
        data_retorno = alerta.data_retorno.strftime("%d/%m/%Y") if alerta.data_retorno else "Sem data"

        paciente = alerta.paciente.nome if alerta.paciente else "-"

        profissional = alerta.profissional.nome if alerta.profissional else "-"

        tipo_retorno = alerta.get_tipo_retorno_display() if alerta.tipo_retorno else "-"

        status = alerta.get_status_display() if alerta.status else "-"

        motivo = alerta.motivo or "-"

        dados.append([
            data_retorno,
            paciente,
            profissional,
            tipo_retorno,
            status,
            motivo,
        ])

    if len(dados) == 1:
        dados.append(["-", "Nenhum alerta encontrado", "-", "-", "-", "-"])

    tabela = Table(
        dados,
        colWidths=[
            2.2 * cm,
            3.5 * cm,
            3.5 * cm,
            3.0 * cm,
            2.3 * cm,
            4.0 * cm,
        ],
        repeatRows=1,
    )

    tabela.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1976d2")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 8),

        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 1), (-1, -1), 7),

        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),

        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#cccccc")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [
            colors.white,
            colors.HexColor("#f5f5f5"),
        ]),

        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))

    elementos.append(tabela)

    doc.build(elementos)

    return response


from django.shortcuts import render
from django.http import JsonResponse
from django.template.loader import render_to_string
from django.db.models import Sum, Q
from datetime import date, timedelta
from .models import Transacao, CategoriaDespesa, Plano
from django.contrib.auth.models import User


def fluxo_caixa(request):
    hoje = date.today()

    # ==========================
    # PEGANDO FILTROS
    # ==========================
    periodo = request.GET.get("periodo")
    busca = request.GET.get("busca")

    tipo = request.GET.getlist("tipo")
    pago = request.GET.get("pago")
    categoria = request.GET.getlist("categoria")
    meio = request.GET.getlist("meio")
    caixa = request.GET.getlist("caixa")
    recibo = request.GET.get("recibo")
    em_atraso = request.GET.get("atraso")

    # Remove valores vazios
    tipo = [t for t in tipo if t]
    categoria = [c for c in categoria if c and c != "todos"]
    meio = [m for m in meio if m and m != "todos"]
    caixa = [c for c in caixa if c and c != "todos"]

    transacoes = Transacao.objects.select_related("categoria", "plano", "profissional").all()

    # ==========================
    # FILTRO DE PERÍODO
    # ==========================
    if periodo == "hoje":
        transacoes = transacoes.filter(data=hoje)

    elif periodo == "30":
        data_inicio = hoje - timedelta(days=30)
        transacoes = transacoes.filter(data__gte=data_inicio)

    elif periodo == "mes":
        transacoes = transacoes.filter(
            data__year=hoje.year,
            data__month=hoje.month
        )

    # ==========================
    # FILTRO BUSCA
    # ==========================
    if busca:
        transacoes = transacoes.filter(
            Q(descricao__icontains=busca) |
            Q(tipo__icontains=busca) |
            Q(valor__icontains=busca)
        )

    # ==========================
    # FILTRO TIPO
    # ==========================
    if tipo:
        transacoes = transacoes.filter(tipo__in=tipo)

    # ==========================
    # FILTRO PAGO
    # ==========================
    if pago in ["0", "1"]:
        transacoes = transacoes.filter(pago=bool(int(pago)))

    # ==========================
    # FILTRO CATEGORIA
    # ==========================
    if categoria:
        transacoes = transacoes.filter(categoria__id__in=categoria)

    # ==========================
    # FILTRO MEIO PAGAMENTO
    # ==========================
    if meio:
        transacoes = transacoes.filter(meio_pagamento__in=meio)

    # ==========================
    # FILTRO CAIXA
    # ==========================
    if caixa:
        transacoes = transacoes.filter(caixa__in=caixa)

    # ==========================
    # FILTRO COM RECIBO
    # ==========================
    if recibo == "1":
        transacoes = transacoes.filter(com_recibo=True)

    # ==========================
    # FILTRO EM ATRASO
    # ==========================
    if em_atraso == "1":
        transacoes = transacoes.filter(
            pago=False,
            data_vencimento__lt=hoje
        )

    # ==========================
    # ORDENAÇÃO
    # ==========================
    transacoes = transacoes.order_by("-data", "-criado_em")

    # ==========================
    # RESUMO
    # ==========================
    
    # RECEITAS
    receitas_recebidas = (
        transacoes.filter(tipo="receita", pago=True)
        .aggregate(Sum("valor"))["valor__sum"] or 0
    )

    receitas_pendentes = (
        transacoes.filter(tipo="receita", pago=False)
        .aggregate(Sum("valor"))["valor__sum"] or 0
    )

    total_receitas_previsto = receitas_recebidas + receitas_pendentes


    # DESPESAS
    despesas_pagas = (
        transacoes.filter(tipo="despesa", pago=True)
        .aggregate(Sum("valor"))["valor__sum"] or 0
    )

    despesas_pendentes = (
        transacoes.filter(tipo="despesa", pago=False)
        .aggregate(Sum("valor"))["valor__sum"] or 0
    )

    total_despesas_previsto = despesas_pagas + despesas_pendentes


    # SALDOS
    saldo = receitas_recebidas - despesas_pagas

    saldo_previsto = (
        total_receitas_previsto -
        total_despesas_previsto
    )


    # CONTADORES
    qtd_receitas = transacoes.filter(tipo="receita").count()

    qtd_despesas = transacoes.filter(tipo="despesa").count()

    qtd_total = transacoes.count()

    categorias = CategoriaDespesa.objects.all().order_by("nome")
    planos = Plano.objects.all().order_by("nome")

    # ✅ lista de profissionais (User) pra usar no select do modal receita
    users = User.objects.all().order_by("username")

    context = {
        "transacoes": transacoes,

        "receitas_recebidas": receitas_recebidas,
        "receitas_pendentes": receitas_pendentes,
        "total_receitas_previsto": total_receitas_previsto,

        "despesas_pagas": despesas_pagas,
        "despesas_pendentes": despesas_pendentes,
        "total_despesas_previsto": total_despesas_previsto,

        "saldo": saldo,
        "saldo_previsto": saldo_previsto,

        "qtd_receitas": qtd_receitas,
        "qtd_despesas": qtd_despesas,
        "qtd_total": qtd_total,

        "categorias": categorias,
        "planos": planos,
        "status_pagamento": Transacao.STATUS_PAGAMENTO,
        "users": users,
    }

    for t in transacoes:
        print(
        "ID:",
        t.id,
        "TIPO:",
        t.tipo,
        "PAGO:",
        t.pago
        )

    if request.headers.get("x-requested-with") == "XMLHttpRequest":
        lista_html = render_to_string(
            "Clinica_Sorriso/financeiro/partials/_lista_transacoes.html",
            context,
            request=request
        )
        resumo_html = render_to_string(
            "Clinica_Sorriso/financeiro/partials/_resumo.html",
            context,
            request=request
        )
        return JsonResponse({
            "lista": lista_html,
            "resumo": resumo_html
        })

    return render(request, "Clinica_Sorriso/financeiro/fluxo_caixa.html", context)


from django.http import JsonResponse
from django.utils.text import slugify
from .models import CategoriaDespesa

def criar_categoria(request):
    if request.method == "POST":
        nome = request.POST.get("nome")

        if not nome:
            return JsonResponse({"erro": "Nome inválido"}, status=400)

        slug = slugify(nome)

        categoria, created = CategoriaDespesa.objects.get_or_create(
            slug=slug,
            defaults={"nome": nome}
        )

        return JsonResponse({
            "id": categoria.id,
            "nome": categoria.nome,
            "slug": categoria.slug
        })

from django.shortcuts import redirect
from decimal import Decimal

def salvar_despesa(request):
    if request.method == "POST":

        descricao = request.POST.get("descricao")
        valor = request.POST.get("valor")
        data_vencimento = request.POST.get("data_vencimento")
        categoria_id = request.POST.get("categoria")
        caixa = request.POST.get("caixa")
        meio_pagamento = request.POST.get("meio_pagamento")
        observacao = request.POST.get("observacao")
        pago = True if request.POST.get("pago") == "on" else False

        Transacao.objects.create(
            descricao=descricao,
            tipo="despesa",
            valor=Decimal(valor),
            data_vencimento=data_vencimento,
            categoria_id=categoria_id if categoria_id else None,
            caixa=caixa.lower(),  # importante
            meio_pagamento=meio_pagamento,
            observacao=observacao,
            pago=pago
        )

        return redirect("fluxo_caixa")

    return redirect("fluxo_caixa")

from decimal import Decimal
from django.shortcuts import redirect
from .models import Transacao

def salvar_receita(request):

    if request.method == "POST":

        paciente = request.POST.get("paciente")
        data_vencimento = request.POST.get("data_vencimento")
        observacao = request.POST.get("observacao")
        plano_id = request.POST.get("plano")
        profissional_id = request.POST.get("profissional")

        tratamentos = request.POST.getlist("tratamento[]")
        valores = request.POST.getlist("valor[]")

        for tratamento, valor in zip(tratamentos, valores):

            # ignora linhas vazias
            if not tratamento or not valor:
                continue

            Transacao.objects.create(

                paciente_nome=paciente,

                descricao=tratamento,

                tipo="receita",

                valor=Decimal(valor),

                data=data_vencimento,

                data_vencimento=data_vencimento,

                observacao=observacao,

                pago=False,

                plano_id=plano_id if plano_id else None,

                profissional_id=profissional_id if profissional_id else None,

            )

    return redirect("fluxo_caixa")

from django.shortcuts import redirect
from django.utils import timezone
from decimal import Decimal
from .models import Transacao, CategoriaDespesa


def criar_despesa(request):
    if request.method == "POST":

        descricao = request.POST.get("descricao")
        valor = request.POST.get("valor")
        data_vencimento = request.POST.get("data_vencimento")
        categoria_id = request.POST.get("categoria")
        caixa = request.POST.get("caixa")
        meio_pagamento = request.POST.get("meio_pagamento")
        observacao = request.POST.get("observacao")
        pago = True if request.POST.get("pago") else False

        categoria = None
        if categoria_id:
            categoria = CategoriaDespesa.objects.get(id=categoria_id)

        Transacao.objects.create(
            descricao=descricao,
            tipo="despesa",
            valor=Decimal(valor),
            data=data_vencimento,  # 👈 mesma data
            data_vencimento=data_vencimento,
            categoria=categoria,
            caixa=caixa,
            meio_pagamento=meio_pagamento,
            observacao=observacao,
            pago=pago,
        )

        return redirect("fluxo_caixa")

    return redirect("fluxo_caixa")

    
import json

def criar_plano(request):
    if request.method == "POST":
        data = json.loads(request.body)
        nome = data.get("nome")

        plano, created = Plano.objects.get_or_create(nome=nome)

        return JsonResponse({
            "id": plano.id,
            "nome": plano.nome
        })
    

from django.http import HttpResponse
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from django.utils.timezone import localdate
from .models import Transacao  # ajuste para seu model real

def exportar_transacoes_pdf(request):

    qs = Transacao.objects.all()

    # ===== mesmos filtros que você já tinha =====

    busca = request.GET.get("busca", "").strip()
    if busca:
        qs = qs.filter(descricao__icontains=busca)

    periodo = request.GET.get("periodo", "")
    hoje = localdate()

    if periodo == "hoje":
        qs = qs.filter(data_vencimento=hoje)
    elif periodo == "mes":
        qs = qs.filter(
            data_vencimento__year=hoje.year,
            data_vencimento__month=hoje.month
        )

    tipos = request.GET.getlist("tipo")
    if tipos:
        qs = qs.filter(tipo__in=tipos)

    pago = request.GET.get("pago")
    if pago in ("0", "1"):
        qs = qs.filter(pago=(pago == "1"))

    # ============================================

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="financeiro.pdf"'

    p = canvas.Canvas(response, pagesize=A4)
    width, height = A4

    y = height - 40

    p.setFont("Helvetica-Bold", 14)
    p.drawString(40, y, "Relatório Financeiro")
    y -= 25

    p.setFont("Helvetica-Bold", 10)
    p.drawString(40, y, "Data")
    p.drawString(100, y, "Descrição")
    p.drawString(350, y, "Tipo")
    p.drawString(420, y, "Valor")
    y -= 15

    p.setFont("Helvetica", 9)

    for t in qs.order_by("-data_vencimento"):

        if y < 50:
            p.showPage()
            y = height - 40
            p.setFont("Helvetica", 9)

        data_str = t.data_vencimento.strftime("%d/%m/%Y") if t.data_vencimento else ""

        p.drawString(40, y, data_str)
        p.drawString(100, y, (t.descricao or "")[:35])
        p.drawString(350, y, str(t.tipo))
        p.drawRightString(520, y, f"R$ {float(t.valor):.2f}".replace(".", ","))

        y -= 14

    p.save()

    return response

# views.py
# views.py
from datetime import timedelta
from django.shortcuts import render
from django.db.models import Sum, DecimalField
from django.db.models.functions import Coalesce
from django.utils.timezone import localdate
from django.contrib.auth.models import User

from .models import Transacao, PagamentoComissao


def comissoes(request):
    ver = request.GET.get("ver", "aberto")  # aberto | historico
    periodo = request.GET.get("periodo", "mes")  # hoje|semana|mes|mes_passado|30|todos|custom
    profissional = request.GET.get("profissional", "")  # id do User (string)
    de = request.GET.get("de", "")  # YYYY-MM-DD
    ate = request.GET.get("ate", "")  # YYYY-MM-DD

    hoje = localdate()

    # =========================
    # HELPER: aplica período em um queryset
    # =========================
    def aplicar_periodo(qs, campo_data: str):
        """
        campo_data: "data" (Transacao) ou "data_pagamento" (PagamentoComissao)
        """
        if periodo == "hoje":
            return qs.filter(**{campo_data: hoje})

        if periodo == "semana":
            inicio = hoje - timedelta(days=hoje.weekday())  # seg
            fim = inicio + timedelta(days=6)                # dom
            return qs.filter(**{f"{campo_data}__range": [inicio, fim]})

        if periodo == "mes":
            return qs.filter(**{f"{campo_data}__year": hoje.year, f"{campo_data}__month": hoje.month})

        if periodo == "mes_passado":
            primeiro_mes_atual = hoje.replace(day=1)
            ultimo_mes_passado = primeiro_mes_atual - timedelta(days=1)
            return qs.filter(**{
                f"{campo_data}__year": ultimo_mes_passado.year,
                f"{campo_data}__month": ultimo_mes_passado.month
            })

        if periodo == "30":
            inicio = hoje - timedelta(days=30)
            return qs.filter(**{f"{campo_data}__gte": inicio, f"{campo_data}__lte": hoje})

        if periodo == "custom":
            if de:
                qs = qs.filter(**{f"{campo_data}__gte": de})
            if ate:
                qs = qs.filter(**{f"{campo_data}__lte": ate})
            return qs

        # "todos" não filtra
        return qs

    # =========================
    # 1) EM ABERTO (igual ao seu, só organizado)
    # =========================
    if ver == "aberto":
        qs = Transacao.objects.filter(tipo="receita", pago=False)

        if profissional:
            qs = qs.filter(profissional_id=profissional)

        qs = aplicar_periodo(qs, "data")

        linhas = (
            qs.values("profissional_id", "profissional__first_name", "profissional__last_name", "profissional__username")
              .annotate(
                  total=Coalesce(
                      Sum("valor"),
                      0,
                      output_field=DecimalField(max_digits=10, decimal_places=2)
                  )
              )
              .order_by("profissional__first_name", "profissional__username")
        )

    # =========================
    # 2) HISTÓRICO (UMA LINHA POR PAGAMENTO)
    # =========================
    else:
        qs = PagamentoComissao.objects.select_related("profissional")

        if profissional:
            qs = qs.filter(profissional_id=profissional)

        qs = aplicar_periodo(qs, "data_pagamento")

        # ✅ aqui NÃO agrupa: uma linha por pagamento
        linhas = qs.order_by("-data_pagamento", "-id")

    # =========================
    # Select de profissionais
    # (inclui quem tem receita OU quem já tem pagamento no histórico)
    # =========================
    ids_receitas = (
        Transacao.objects.filter(tipo="receita")
        .exclude(profissional__isnull=True)
        .values_list("profissional_id", flat=True)
        .distinct()
    )

    ids_pagamentos = (
        PagamentoComissao.objects
        .values_list("profissional_id", flat=True)
        .distinct()
    )

    profissionais = User.objects.filter(id__in=set(ids_receitas) | set(ids_pagamentos)).order_by("first_name", "username")

    context = {
        "ver": ver,
        "periodo": periodo,
        "profissional": profissional,
        "de": de,
        "ate": ate,
        "profissionais": profissionais,
        "linhas": linhas,
    }
    return render(request, "Clinica_Sorriso/financeiro/comissoes.html", context)

from django.http import HttpResponse
from django.contrib.auth.models import User
from django.db.models import Sum
from django.db.models.functions import Coalesce
from django.db.models import DecimalField, Value
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from django.utils.timezone import localdate
from datetime import timedelta
from .models import Transacao

def exportar_comissoes_pdf(request):
    ver = request.GET.get("ver", "aberto")             # aberto | historico
    periodo = request.GET.get("periodo", "mes")        # hoje | semana | mes | mes_passado | 30 | todos | custom
    profissional = request.GET.get("profissional", "") # id do User
    de = request.GET.get("de", "")
    ate = request.GET.get("ate", "")

    qs = Transacao.objects.filter(tipo="receita")

    # filtra por profissional (FK)
    if profissional:
        qs = qs.filter(profissional_id=profissional)

    hoje = localdate()

    # período
    if periodo == "hoje":
        qs = qs.filter(data_vencimento=hoje)
    elif periodo == "semana":
        ini = hoje - timedelta(days=hoje.weekday())
        fim = ini + timedelta(days=6)
        qs = qs.filter(data_vencimento__range=[ini, fim])
    elif periodo == "mes":
        qs = qs.filter(data_vencimento__year=hoje.year, data_vencimento__month=hoje.month)
    elif periodo == "mes_passado":
        ano = hoje.year
        mes = hoje.month - 1
        if mes == 0:
            mes = 12
            ano -= 1
        qs = qs.filter(data_vencimento__year=ano, data_vencimento__month=mes)
    elif periodo == "30":
        qs = qs.filter(data_vencimento__gte=hoje - timedelta(days=30))
    elif periodo == "custom":
        if de and ate:
            qs = qs.filter(data_vencimento__range=[de, ate])
        elif de:
            qs = qs.filter(data_vencimento__gte=de)
        elif ate:
            qs = qs.filter(data_vencimento__lte=ate)

    # status (em aberto vs histórico)
    if ver == "aberto":
        qs = qs.filter(pago=False)
    else:
        qs = qs.filter(pago=True)

    # agrupa por profissional
    linhas = (
        qs.values("profissional_id", "profissional__first_name", "profissional__last_name", "profissional__username")
          .annotate(
              total=Coalesce(
                  Sum("valor"),
                  Value(0, output_field=DecimalField(max_digits=10, decimal_places=2))
              )
          )
          .order_by("profissional__first_name", "profissional__username")
    )

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="comissoes.pdf"'

    p = canvas.Canvas(response, pagesize=A4)
    width, height = A4
    y = height - 40

    p.setFont("Helvetica-Bold", 14)
    p.drawString(40, y, "Relatório de Comissões")
    y -= 18

    p.setFont("Helvetica", 10)
    p.drawString(40, y, f"Ver: {ver} | Período: {periodo}")
    y -= 22

    p.setFont("Helvetica-Bold", 10)
    p.drawString(40, y, "Profissional")
    p.drawRightString(520, y, "Total")
    y -= 14

    p.setFont("Helvetica", 10)

    for row in linhas:
        if y < 60:
            p.showPage()
            y = height - 40
            p.setFont("Helvetica", 10)

        nome = (row["profissional__first_name"] or "").strip()
        sobrenome = (row["profissional__last_name"] or "").strip()
        usern = row["profissional__username"] or "—"
        display = (f"{nome} {sobrenome}".strip()) or usern

        total = row["total"] or 0
        p.drawString(40, y, display[:45])
        p.drawRightString(520, y, f"R$ {float(total):.2f}".replace(".", ","))
        y -= 16

    p.save()
    return response

from django.shortcuts import get_object_or_404, redirect

def receber_transacao(request, pk):

    print("================================")
    print("ENTROU NA VIEW RECEBER")
    print("METHOD =", request.method)
    print("POST =", request.POST)
    print("PK =", pk)
    print("================================")

    if request.method == "POST":

        transacao = get_object_or_404(Transacao, pk=pk)

        print("ANTES:", transacao.pago)

        valor = request.POST.get("valor_pago")

        if valor:
            transacao.valor = valor.replace(",", ".")

        transacao.meio_pagamento = request.POST.get("meio")
        transacao.data = request.POST.get("data_pagamento")
        transacao.observacao = request.POST.get("observacao")
        transacao.pago = True

        transacao.save()

        print("DEPOIS:", transacao.pago)

    return redirect("fluxo_caixa")


from django.shortcuts import get_object_or_404, redirect
from decimal import Decimal
from .models import Transacao

def pagar_despesa(request, transacao_id):

    print("================================")
    print("ENTROU NA VIEW PAGAR")
    print("METHOD =", request.method)
    print("POST =", request.POST)
    print("================================")

    if request.method == "POST":

        transacao = get_object_or_404(
            Transacao,
            id=transacao_id
        )

        # Marca como paga
        transacao.pago = True

        # Atualiza os dados vindos do modal
        valor_pago = request.POST.get("valor_pago")

        if valor_pago:
            valor_pago = valor_pago.replace(",", ".")
            transacao.valor = Decimal(valor_pago)

        transacao.meio_pagamento = request.POST.get(
            "meio_pagamento",
            transacao.meio_pagamento
        )

        transacao.observacao = request.POST.get(
            "observacao",
            transacao.observacao
        )

        data_pagamento = request.POST.get("data_pagamento")
        if data_pagamento:
            transacao.data = data_pagamento

        transacao.save()

        print("SALVOU COM SUCESSO")

    return redirect("fluxo_caixa")

from django.shortcuts import get_object_or_404, redirect
from django.contrib import messages
from .models import Transacao

def excluir_despesa(request, transacao_id):
    # Busca a transação ou retorna erro 404 se não existir
    transacao = get_object_or_404(Transacao, id=transacao_id)
    
    # Executa a exclusão no banco de dados
    transacao.delete()
    
    # Opcional: Adiciona uma mensagem de sucesso que você pode exibir na tela
    messages.success(request, "Despesa excluída com sucesso!")
    
    return redirect("fluxo_caixa")

def editar_despesa(request, transacao_id):

    print("POST =", request.POST)

    if request.method == "POST":

        transacao = get_object_or_404(
            Transacao,
            id=transacao_id
        )

        transacao.descricao = request.POST.get("descricao")

        valor = request.POST.get("valor")
        if valor:
            transacao.valor = valor.replace(",", ".")

        transacao.meio_pagamento = request.POST.get("meio_pagamento")

        transacao.observacao = request.POST.get("observacao")

        transacao.save()

        print("SALVOU")

    return redirect("fluxo_caixa")

def cancelar_pagamento(request, transacao_id):

    transacao = get_object_or_404(
        Transacao,
        id=transacao_id
    )

    transacao.pago = False

    transacao.save()

    return redirect("fluxo_caixa")

from django.shortcuts import get_object_or_404, redirect
from .models import Transacao

def cancelar_recebimento(request, transacao_id):

    transacao = get_object_or_404(
        Transacao,
        id=transacao_id
    )

    transacao.pago = False

    transacao.meio_pagamento = None

    transacao.save()

    return redirect("fluxo_caixa")

from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from .models import Etiqueta # Substitua pelo caminho real do seu model de Etiqueta

def salvar_etiqueta_ajax(request):
    if request.method == "POST":
        nome = request.POST.get("nome", "").strip()
        cor = request.POST.get("cor", "#28a745").strip()

        if not nome:
            return JsonResponse({"success": False, "error": "O nome da etiqueta é obrigatório."}, status=400)

        # Cria a etiqueta no banco de dados
        etiqueta = Etiqueta.objects.create(nome=nome, cor=cor)

        return JsonResponse({
            "success": True,
            "id": etiqueta.id,
            "nome": etiqueta.nome,
            "cor": etiqueta.cor
        })
    
    return JsonResponse({"success": False, "error": "Método não permitido."}, status=405)

# =========================================================================
# NOVA VIEW: RESPONSÁVEL POR CRIAR A ETIQUETA VIA REQUISIÇÃO AJAX (JAVASCRIPT)
# =========================================================================
def salvar_etiqueta_ajax(request):
    if request.method == "POST":
        # Converte o nome para minúsculo ou limpa os espaços para evitar duplicados bobos
        nome = request.POST.get("nome", "").strip()
        cor = request.POST.get("cor", "#28a745").strip()

        if not nome:
            return JsonResponse({"success": False, "error": "O nome da etiqueta é obrigatório."}, status=400)

        # get_or_create tenta buscar; se não achar pelo nome, ele cria uma nova!
        etiqueta, criado = Etiqueta.objects.get_or_create(
            nome__iexact=nome, # Busca ignorando maiúsculas/minúsculas
            defaults={'nome': nome, 'cor': cor}
        )

        return JsonResponse({
            "success": True,
            "id": etiqueta.id,
            "nome": etiqueta.nome,
            "cor": etiqueta.cor,
            "criado_agora": criado # Informação extra caso precise
        })
    
    return JsonResponse({"success": False, "error": "Método não permitido."}, status=405)

import json
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_protect
from django.apps import apps

@csrf_protect
def pagar_despesas_em_lote(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            ids = data.get('ids', [])
            data_pagamento = data.get('data_pagamento')
            meio_pagamento = data.get('meio_pagamento')
            observacao = data.get('observacao', '')

            # Carrega o modelo financeiro dinamicamente para evitar ImportError
            # Se a classe se chamar 'Despesas' ou 'FluxoCaixa', mude apenas o nome abaixo
            ModelFinanceiro = apps.get_model('Clinica_Sorriso', 'Despesa')

            for despesa_id in ids:
                despesa = ModelFinanceiro.objects.get(id=despesa_id)
                
                # ATENÇÃO: Copie exatamente os mesmos campos que sua segunda tela altera!
                # Exemplos comuns de campos:
                despesa.status = 'pago'  # ou despesa.pago = True
                despesa.data_pagamento = data_pagamento
                despesa.meio_pagamento = meio_pagamento
                
                if observacao:
                    despesa.observacao = f"{getattr(despesa, 'observacao', '')}\n{observacao}".strip()
                
                despesa.save() # Dispara a atualização de saldos

            return JsonResponse({'sucesso': True})
            
        except Exception as e:
            return JsonResponse({'sucesso': False, 'erro': str(e)})

    return JsonResponse({'sucesso': False, 'erro': 'Método inválido'})

from decimal import Decimal
from django.http import JsonResponse
from django.shortcuts import get_object_or_404
from django.views.decorators.csrf import csrf_protect
from .models import Transacao

@csrf_protect
def editar_receita(request, receita_id):
    if request.method == 'POST':
        try:
            # 1. Busca a receita exata no banco de dados pelo ID
            receita = get_object_or_404(Transacao, id=receita_id)
            
            # 2. Captura os valores enviados pelo JavaScript
            descricao = request.POST.get('descricao')
            valor_texto = request.POST.get('valor', '0')
            data_vencimento = request.POST.get('data_vencimento')
            observacao = request.POST.get('observacao', '')
            paciente_nome = request.POST.get('paciente_nome')
            plano_id = request.POST.get('plano')
            profissional_id = request.POST.get('dentista')

            # 3. CONVERSÃO DIRETA E LIMPA (O JavaScript já mandou o número com ponto exato)
            # Removemos os replaces antigos que multiplicavam o número por 100
            valor_decimal = Decimal(valor_texto) if valor_texto else Decimal('0.00')

            # 4. Injeta as alterações diretamente nas colunas da tabela Transacao
            receita.paciente_nome = paciente_nome
            receita.descricao = descricao
            receita.valor = valor_decimal
            
            # Força a gravação direta da data para atualizar os filtros
            receita.data = data_vencimento
            receita.data_vencimento = data_vencimento
                
            receita.observacao = observacao
            
            # Vincula as chaves estrangeiras tratando de forma segura
            receita.plano_id = int(plano_id) if plano_id and plano_id.isdigit() else None
            receita.profissional_id = int(profissional_id) if profissional_id and profissional_id.isdigit() else None

            # 5. Salva permanentemente no banco de dados
            receita.save()

            print("========== RECEITA ==========")
            print(receita.id)
            print(descricao)
            print(valor_decimal)
            print(data_vencimento)
            print(observacao)
            print("=============================")

            # Retorna o sinalizador de sucesso para o JavaScript recarregar a página
            return JsonResponse({'sucesso': True})
            
        except Exception as e:
            return JsonResponse({'sucesso': False, 'erro': str(e)})

    return JsonResponse({'sucesso': False, 'erro': 'Método inválido'})

